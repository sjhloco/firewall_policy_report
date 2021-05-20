#!/usr/bin/env python

import argparse
from getpass import getpass
import os
from datetime import date
from collections import defaultdict
import yaml
from rich.console import Console
from rich.theme import Theme
from rich.progress import track
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, colors, PatternFill, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule


######################## Variables to change dependant on environment ########################
# Directory and filename where saves MS and device prefixes to. By default is the users home directory
directory = os.path.dirname(__file__)
report_name = 'ACLreport_' + date.today().strftime('%Y%m%d')
input_file = 'input.yml'
# Header names and columns widths for the XL sheet
header = {'Policy/ACL Name':25, 'Line Number':17, 'Access':18, 'Protocol':12, 'Source Address':23, 'Source Service':14, 'Destination Address':23,
          'Destination Service':26, 'Hit Count':14, 'Date Last Hit':17, 'Time Last Hit':17, 'State':10}


################################## Multi-Use functions ##################################
# INPUT_VAL: Validates the input file has the correct dictionaires and format for each FW type (ASA or CKP)
def input_val(my_vars, fw_type, errors):
    if my_vars.get(fw_type) != None:
        try:
            # Assert if the FW dictionary exists
            assert my_vars[fw_type].get('fw') != None, ":x: [b red]Error[/b red] - [i cyan]'{}'[/i cyan] must have a [black]'fw'"\
                                                        "[/black] dictionary containing a list of firewalls".format(fw_type)
            # Assert that the FW dictionary is a list
            assert isinstance(my_vars[fw_type].get('fw'), list), ":x: [b red]Error[/b red] - [i cyan]'{}'[/i cyan] must have a "\
                                                        "[black]'fw'[/black] dictionary containing a list of firewalls".format(fw_type)
            for each_fw in my_vars[fw_type]['fw']:
                # Assert that each FW has a ip_name
                assert each_fw.get('ip_name') != None, ":x: [b red]Error[/b red] - Every [i cyan]'{}'[/i cyan] [black]'fw'[/black] object " \
                                              "must have an [black]'ip_name'[/black] dictionary of hostname or IP address".format(fw_type)
                # Assert that if the global or FW_type username/password is not defined the individual FW object username/password is
                user = each_fw.get('user', my_vars[fw_type].get('user', my_vars.get('user', None)))
                pword = each_fw.get('pword', my_vars[fw_type].get('pword', my_vars.get('pword', None)))
                if user == None and pword == None:
                    errors.append(":x: [b red]Error[/b red] - The [i cyan]'{}'[/i cyan] fw [black]{}[black] has no username or password".format(
                                  fw_type, each_fw['ip_name']))
                elif user == None:
                    errors.append(":x: [b red]Error[/b red] - The [i cyan]'{}'[/i cyan] fw [black]{}[black] has no username".format(fw_type, each_fw['ip_name']))
                elif pword == None:
                    errors.append(":x: [b red]Error[/b red] - The [i cyan]'{}'[/i cyan] fw [black]{}[black] has no password".format(fw_type, each_fw['ip_name']))
        except Exception as e:
            errors.append(e)

# FW_DICT: Creates a list of firewalls under a dictionary of the FW type (ASA or CKP)
def create_fw_dict(my_vars, fw_type):
    all_fw = defaultdict(list)
    try:
        for each_fw in my_vars[fw_type]['fw']:
            # Gets device_type creds, if none exist uses global creds
            dvc_type_creds = dict(user=my_vars[fw_type].get('user', my_vars.get('user')), pword=my_vars[fw_type].get('pword', my_vars.get('pword')))
            # Adds dict to asa dict list in format {fw_ip: (user, pword)} replacing global user/pword if more specific set for the device
            all_fw[fw_type].append({each_fw['ip_name']: (each_fw.get('user', dvc_type_creds['user']), each_fw.get('pword', dvc_type_creds['pword']))})
        return all_fw
    except Exception as e:
        print('Could not create the {} FW list {} because of the error:'.format(fw_type, e))


################################## 1. User input collected ##################################
# Optional flags user can enter to customize what is run, if nothing is entered uses the default options
def create_parser():
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--input', default=input_file, help='Input file holding FW IPs and credentials (default: %(default)s)')
    parser.add_argument('-u', '--user', help='Global username for all devices, overides that set in input file')
    parser.add_argument('-l', '--location', default=directory, help='Location to save and run the report (default: %(default)s)')
    parser.add_argument('-n', '--name', default=report_name, help='Name for the report (default: %(default)s)')
    return vars(parser.parse_args())


################################## 2. Validation and login credentials ##################################
def validate_creds(args, fw_types):
    location_exist = "no"
    file_exist = "yes"
    all_fw = {}
    errors = []
    print('Checking the input file and options entered are valid...')

    #2a. LOCATION: If the directory location does not exist ask whether to create it or if not exit
    if not os.path.exists(args['location']):
        rc.print("The directory [i cyan]'{}'[/i cyan] does not exist, do you want to create it?".format(args['location']))
        while location_exist == "no":
            answer = rc.input('[b green3]y or n: [/b green3]').lower()
            if answer == 'y':
                os.makedirs(args['location'])
            elif answer == 'n':
                print('Please rerun the script with the correct directory')
                exit()
            else:
                rc.print(":x: [b red]Error[/b red] - The only acceptable options are 'y' or 'n'")
        else:
            location_exist = "yes"

    #2b. NAME: If the file already exists user is asked whether they wish to overwrite it
    while file_exist == "yes":
        if os.path.exists(os.path.join(args['location'], args['name'] + ".xlsx")):
            rc.print("The output file [i cyan]{}[/i cyan] already exist, do you want to overwrite it?".format(args['name'] + ".xlsx"))
            answer = rc.input('[b green3]y or n: [/b green3]').lower()
            if answer == 'n':
                args['name'] = input("Please enter a new name for the output file: ")
            elif answer == 'y':
                file_exist = "no"
            else:
                rc.print(":x: [b red]Error[/b red] - The only acceptable options are 'y' or 'n'")
        else:
            file_exist = "no"

    # 2c. LOAD: Load the input file into a dictionary
    with open(os.path.join(args['location'], args['input']), 'r') as file_content:
        my_vars = yaml.load(file_content, Loader=yaml.FullLoader)
    # 2d. CREDS: If user defined in args asks for password and overwrites global creds in input file
    if args['user'] != None:
        my_vars['user'] = args['user']
        my_vars['pword'] = getpass("Enter global fw password for user '{}': ".format(args['user']))
    #2e. INPUT_VAL: Validates input dictionaries, if any errors exits script
    for each_dvc in fw_types:
        input_val(my_vars, each_dvc, errors)
    if len(errors) != 0:
        for each_err in errors:
            rc.print(str(each_err))
        exit()

    # 2f. FW_DICT: Creates a list of firewalls under a dictionary of the FW type (ASA or CKP) in format {fw_type: [{fw_ip: (user, pword)}]}
    for each_dvc in fw_types:
        if my_vars.get(each_dvc) != None:
            all_fw.update(create_fw_dict(my_vars, each_dvc))
    return all_fw


###################################### 3. Logon ######################################
# Logon to devices and get Session ID back which is then used for any subsequent connections
def logon(fw_types, fw_cred):
    fw_sid = defaultdict(dict)
    import_fw = {}
    errors = []

    # FW_SID: Creates nested dict is in the format {asa: {asa1_ip: asa1_sid, asa_ip2: asa2_sid}, ckp: {ckp1_ip: ckp1_sid, ckp2_ip: ckp2_sid}}
    for each_type in fw_types:
        if fw_cred.get(each_type) != None:
            # Import FW type module as a dynamic variable, this allows the module to be specified using string from 'fw_types' list
            import_fw.update({each_type: __import__(each_type)})
            # for each_fw in fw_cred[each_type]:
            for each_fw in track(fw_cred[each_type], 'Testing ' + each_type + ' username/password and device connectivity'):
                dev_ip = list(each_fw.keys())[0]
                fw_sid[each_type][dev_ip] = import_fw[each_type].login(dev_ip, list(each_fw.values())[0][0], list(each_fw.values())[0][1])

    # FAILFAST: Any sessions returning an error (False) print errors, close all other open sessions and exit. If have a SID (True) change tuple to just SID
    for each_type, each_fw_sid in fw_sid.items():
        for fw, sid in (list(each_fw_sid.items())):
            if sid[0] == True:
                fw_sid[each_type][fw] = sid[1]
            elif sid[0] == False:
                rc.print(sid[1])
                errors.append(fw)
    if len(errors) != 0:
        rc.print(':x: [b red]Error[/b red] - Missing session IDs for [i]{}[/i], closing ALL connections.'.format(str(errors).replace('[', '').replace(']', '')))
        logoff(import_fw, fw_sid)
    return import_fw, fw_sid


############################## Logoff - Gracefully close all device conns ###################################
def logoff(import_fw, fw_sid):
        for fw_type, fw_sid in fw_sid.items():
            for fw, sid in fw_sid.items():
                if not isinstance(sid, tuple):
                    import_fw[fw_type].logoff(fw, sid)
        exit()


############################ Toggle colour - alternative colour at each loop iteration ###########################
def toggle_colour(last=[0]):
    colours = ['green4', 'dark_sea_green4']
    colour = colours[last[0]]
    last[0] = (last[0] + 1) % 2         # ensure the index is 0 or 1 alternatively
    return colour


 ################################## 5. Build XL worksheet ##################################
def create_xls(args, acl):
    print('Creating the spreadsheet...')
    filename = os.path.join(args['location'], args['name'] + ".xlsx")

    # 5a. Create a workbook per device
    wb = Workbook()
    for dvc, dvc_acl in acl.items():
        ws1 = wb.create_sheet(title=dvc)

        # 5b. Add the headers, set font, colour and column width (from header dictionary)
        for col, head in zip(range(1,len(header) + 1), header.items()):
            ws1['{}1'.format(get_column_letter(col))] = head[0]      # get_column_letter converts number to letter
            ws1['{}1'.format(get_column_letter(col))].fill = PatternFill(bgColor=colors.Color("00DCDCDC"))
            ws1['{}1'.format(get_column_letter(col))].font = Font(bold=True, size=14)
            ws1.column_dimensions[get_column_letter(col)].width = head[1]
        # 5c. Add the ACE entries. The columns holding numbers are changed to integers
        for ace in dvc_acl:
            ace[1] = int(ace[1])
            ace[8] = int(ace[8])
            if ace[7].isdigit():
                ace[7] = int(ace[7])
            ws1.append(ace)
        # 5d. Required as making ace[7] messes up alignment for that column
        for col in ws1.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='left')
        # 5e. Add a key at start with info on the colourised rows for ACEs with frequent hit-cnts
        ws1.insert_rows(1)
        ws1.insert_rows(2)
        keys = {'A1': 'Key:', 'B1':'Hit in last 1 day', 'E1':'Hit in last 7 days', 'G1':'Hit in last 30 days', 'I1':'Inactive'}
        colour  = {'B1':'E6B0AA', 'E1':'A9CCE3', 'G1':'F5CBA7', 'I1':'D4EFDF'}

        for cell, val in keys.items():
            ws1[cell] = val
        ws1['A1'].font = Font(bold=True)
        for cell, col in colour.items():
            ws1[cell].fill = PatternFill(start_color=col, end_color=col, fill_type='solid')

        ws1.freeze_panes = ws1['A4']                    # Freezes the top row (A1) so remains when scrolling
        ws1.auto_filter.ref = 'A3:L4'                   # Adds dropdown to headers to the headers

        # 5f. Colours used for columns dependant on the last hit data (J column). Formula is a standard XL formula
        style_grn = DifferentialStyle(fill=PatternFill(bgColor=colors.Color("00D4EFDF")))
        rule_inactive = Rule(type="expression",formula=['=$L1="inactive"'], dxf=style_grn)
        style_red = DifferentialStyle(fill=PatternFill(bgColor=colors.Color("00E6B0AA")))
        rule_1day = Rule(type="expression",formula=["=AND(TODAY()-$J1>=0,TODAY()-$J1<=1)"], dxf=style_red)
        style_blu = DifferentialStyle(fill=PatternFill(bgColor=colors.Color("00A9CCE3")))
        rule_7day = Rule(type="expression", formula=["=AND(TODAY()-$J1>=0,TODAY()-$J1<=7)"], dxf=style_blu)
        style_org = DifferentialStyle(fill=PatternFill(bgColor=colors.Color("00F5CBA7")))
        rule_30day = Rule(type="expression", formula=["=AND(TODAY()-$J1>=0,TODAY()-$J1<=30)"], dxf=style_org)

        # 5g. Apply the rules to workbook and save it
        for rule in [rule_inactive, rule_1day, rule_7day, rule_30day]:
            ws1.conditional_formatting.add(ws1.dimensions, rule)

    wb.remove(wb['Sheet'])
    wb.save(filename)
    rc.print(':white_heavy_check_mark: Firewall policy report [b blue]{}[/b blue] has been created'.format(filename))


###################################### Run the scripts ######################################
def main():
    global rc
    rc = Console(theme=Theme({"repr.str": "black", "repr.ipv4": "black", "repr.number": "black"}))
    rc.print('\n' + '=' * 30, '[b purple4]Firewall Policy Report v0.1[/b purple4]', '=' * 30)

    # Device types to loop through, match name of the device type python files (asa.py, ckp.py)
    fw_types = ['asa', 'ckp']

	# 1. Gather input from user
    args = create_parser()
    # 2. Validate location and filename and create list of FWs
    fw_cred = validate_creds(args, fw_types)

    # 3. Check login details and create a nested dictionary of sessions for each device
    if len(fw_cred) != None:
        import_fw, fw_sid = logon(fw_types, fw_cred)

    # 4. Gather ACLs from devices then format the data to create new data-models of {fwip_acl: [non_expanded_acl], fw_ip_exp_acl: [expanded_acl]}
    acl = {}
    for fw_type, details in fw_sid.items():
        for fw, sid in details.items():
            colour = toggle_colour()
            rc.print('[{}]Gathering and formatting ACL information from the {} [i]{}[/i], be patient it can take a while...[/{}]'.format(colour, fw_type, fw, colour))
            acl_brief, acl_expanded = import_fw[fw_type].get_acls(fw, sid)

            import_fw[fw_type].format_acl(fw, acl_brief, acl_expanded)
            acl.update(import_fw[fw_type].format_acl(fw, acl_brief, acl_expanded))

    # 5. Build the Excel worksheet, a separate sheet per device
    create_xls(args, acl)

    #6. Logoff sessions form all firewalls
    logoff(import_fw, fw_sid)

if __name__ == '__main__':
    main()
